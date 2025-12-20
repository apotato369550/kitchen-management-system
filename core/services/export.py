"""
Export service for generating PDF and Excel files from kitchen management data.
"""
from datetime import datetime
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def get_export_filename(module_name: str, export_format: str) -> str:
    """Generate timestamped filename for exports."""
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    return f"{module_name}_{export_format}_{timestamp}"


def export_to_excel(data_dict: dict, headers: list) -> BytesIO:
    """
    Generate Excel workbook from data dictionary.

    Args:
        data_dict: Dictionary with 'data' (list of dicts) and optional 'title', 'summary'
        headers: List of column headers

    Returns:
        BytesIO object containing Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = data_dict.get('sheet_name', 'Data')

    # Add title if provided
    if data_dict.get('title'):
        ws['A1'] = data_dict['title']
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:Z1')
        current_row = 3
    else:
        current_row = 1

    # Add export date
    ws[f'A{current_row}'] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws[f'A{current_row}'].font = Font(italic=True, size=9)
    current_row += 2

    # Add headers
    header_row = current_row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data rows
    data_rows = data_dict.get('data', [])
    for row_idx, row_data in enumerate(data_rows, header_row + 1):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = row_data.get(header, '')
            cell.alignment = Alignment(horizontal="left", vertical="center")

            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border

    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=header_row, max_row=header_row + len(data_rows)):
            try:
                if len(str(row[col_idx - 1].value)) > max_length:
                    max_length = len(str(row[col_idx - 1].value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Add summary if provided
    if data_dict.get('summary'):
        summary_row = header_row + len(data_rows) + 2
        ws[f'A{summary_row}'] = "Summary"
        ws[f'A{summary_row}'].font = Font(bold=True, size=11)

        for idx, (key, value) in enumerate(data_dict['summary'].items(), 1):
            ws[f'A{summary_row + idx}'] = f"{key}:"
            ws[f'B{summary_row + idx}'] = value
            ws[f'A{summary_row + idx}'].font = Font(bold=True)

    # Return BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_to_pdf(data_dict: dict, headers: list) -> BytesIO:
    """
    Generate PDF from data dictionary.

    Args:
        data_dict: Dictionary with 'data' (list of dicts), 'title', 'summary'
        headers: List of column headers

    Returns:
        BytesIO object containing PDF file
    """
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter)

    styles = getSampleStyleSheet()
    story = []

    # Title
    if data_dict.get('title'):
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12,
            alignment=1  # center
        )
        story.append(Paragraph(data_dict['title'], title_style))

    # Export date
    date_text = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    story.append(Paragraph(date_text, styles['Normal']))
    story.append(Spacer(1, 0.3 * inch))

    # Prepare table data
    table_data = [headers]
    for row_data in data_dict.get('data', []):
        row = [str(row_data.get(header, '')) for header in headers]
        table_data.append(row)

    # Create table
    col_widths = [2.5 * inch / len(headers)] * len(headers)
    table = Table(table_data, colWidths=col_widths)

    # Style table
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

        # Data styling
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ]))

    story.append(table)

    # Add summary if provided
    if data_dict.get('summary'):
        story.append(Spacer(1, 0.3 * inch))
        story.append(Paragraph("Summary", styles['Heading2']))

        for key, value in data_dict['summary'].items():
            summary_text = f"<b>{key}:</b> {value}"
            story.append(Paragraph(summary_text, styles['Normal']))

    # Add footer
    story.append(Spacer(1, 0.3 * inch))
    footer_text = "Cebu Best Value Trading - Kitchen Management System"
    story.append(Paragraph(footer_text, styles['Normal']))

    # Build PDF
    doc.build(story)
    output.seek(0)
    return output
